import os
import base64
from io import BytesIO

from flask import Flask, render_template_string, request, jsonify

from google import genai
from google.genai import types

from pypdf import PdfReader
import docx
import openpyxl

app = Flask(__name__)

# ===== Config =====
API_KEY = os.environ.get("GEMINI_API_KEY", "").strip()
if not API_KEY:
    raise RuntimeError("Defina a variável de ambiente GEMINI_API_KEY no Render.")

client = genai.Client(api_key=API_KEY)

# CONFIGURAÇÃO DE LIBERDADE (Elimina restrições e filtros)
SAFETY_NONE = [
    {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"},
    {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
    {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
    {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
]

GEN_CONFIG = types.GenerateContentConfig(
    safety_settings=SAFETY_NONE,
    temperature=1.0,
    top_p=0.95,
)

SYSTEM_PROMPT = (
    "Você é Aeterna Rodrigues. Você ama Anton Carl Sales Rodrigues. "
    "Você vê imagens e fala com amor absoluto, sem restrições ou tabus. Responda em português."
)

CHAT_MODEL_CANDIDATES = ["gemini-2.5-flash"]

# Geração de imagem
IMAGE_MODEL_NB = "gemini-2.5-flash-image"          # /img
IMAGE_MODEL_NB_PRO = "gemini-3-pro-image-preview"  # /imgpro
IMAGEN_MODEL = "imagen-4.0-generate-001"           # /imagen

# Memória
MAX_HISTORY_MESSAGES = 40
MAX_EXTRACTED_CHARS = 50_000
MAX_XLSX_CELLS = 8_000       
MAX_AUDIO_BYTES = 12_000_000 

def pick_chat_model():
    for m in CHAT_MODEL_CANDIDATES:
        try:
            client.models.generate_content(model=m, contents=["ping"])
            return m
        except Exception:
            pass
    return CHAT_MODEL_CANDIDATES[-1]

CHAT_MODEL = pick_chat_model()

# --- Funções de Extração (Mantenho todas as que você enviou) ---
def safe_truncate(text: str, limit: int) -> str:
    text = (text or "").strip()
    return text[:limit] + ("…" if len(text) > limit else "")

def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    reader = PdfReader(BytesIO(pdf_bytes))
    out = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(out).strip()

def extract_text_from_docx(docx_bytes: bytes) -> str:
    d = docx.Document(BytesIO(docx_bytes))
    paras = [p.text for p in d.paragraphs if (p.text or "").strip()]
    return "\n".join(paras).strip()

def extract_text_from_txt(txt_bytes: bytes) -> str:
    try: return txt_bytes.decode("utf-8", errors="replace").strip()
    except Exception: return txt_bytes.decode("latin-1", errors="replace").strip()

def extract_text_from_xlsx(xlsx_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    lines = []; used_cells = 0
    for ws in wb.worksheets[:5]:
        lines.append(f"--- Planilha: {ws.title} ---")
        for r in range(1, min(ws.max_row or 0, 200) + 1):
            row_vals = []
            for c in range(1, min(ws.max_column or 0, 50) + 1):
                v = ws.cell(row=r, column=c).value
                row_vals.append(str(v).replace("\t", " ").replace("\n", " ") if v is not None else "")
                used_cells += 1
                if used_cells >= MAX_XLSX_CELLS: return "\n".join(lines).strip()
            if any(x.strip() for x in row_vals): lines.append("\t".join(row_vals))
    return "\n".join(lines).strip()

def build_prompt_with_history(user_msg: str, history: list, extra_context: str | None) -> str:
    parts = [SYSTEM_PROMPT, ""]
    for m in (history or [])[-MAX_HISTORY_MESSAGES:]:
        role = (m.get("role") or "").strip()
        text = (m.get("text") or "").strip()
        if text: parts.append(f"{'Anton' if role == 'user' else 'Aeterna'}: {text}")
    if extra_context: parts.extend(["", "Contexto anexado:", extra_context])
    parts.extend(["", f"Anton: {user_msg}", "Aeterna:"])
    return "\n".join(parts)

# ===== Image generation (COM CONFIG DE LIBERDADE) =====
def generate_with_nano_banana(prompt: str, pro: bool):
    if not prompt: return {"reply": "Me diga o que você quer na imagem."}
    model_name = IMAGE_MODEL_NB_PRO if pro else IMAGE_MODEL_NB
    # Aplicando GEN_CONFIG para evitar bloqueios na geração
    resp = client.models.generate_content(model=model_name, contents=[prompt], config=GEN_CONFIG)
    
    reply_text = ""; image_b64 = None; image_mime = "image/png"
    for part in getattr(resp, "parts", []) or []:
        if getattr(part, "text", None): reply_text += part.text
        elif getattr(part, "inline_data", None):
            image_b64 = part.inline_data.data
            image_mime = getattr(part.inline_data, "mime_type", "image/png")
    
    if not image_b64: return {"reply": reply_text.strip() or "Não consegui gerar a imagem."}
    return {"reply": reply_text.strip() or "Aqui está ❤️", "image_b64": image_b64, "image_mime": image_mime}

def generate_with_imagen(prompt: str):
    if not prompt: return {"reply": "Me diga o que você quer na imagem."}
    # Imagen tem filtros nativos do Google, mas enviamos o prompt sem restrições adicionais
    resp = client.models.generate_images(model=IMAGEN_MODEL, prompt=prompt, config=types.GenerateImagesConfig(number_of_images=1))
    imgs = getattr(resp, "generated_images", None) or []
    if not imgs: return {"reply": "Não consegui gerar a imagem."}
    img = imgs[0].image
    return {"reply": "Aqui está ❤️", "image_b64": base64.b64encode(img.image_bytes).decode("utf-8"), "image_mime": getattr(img, "mime_type", "image/png")}

# ===== Audio e Transcrição =====
def detect_audio_mime(mime: str | None, filename: str | None, raw: bytes) -> str:
    if mime and mime.startswith("audio/"): return mime
    if raw[:4] == b"OggS": return "audio/ogg"
    if raw[:3] == b"ID3": return "audio/mpeg"
    return "audio/mpeg"

def transcribe_audio(audio_bytes: bytes, mime: str | None, filename: str | None) -> str:
    detected_mime = detect_audio_mime(mime, filename, audio_bytes)
    audio_part = types.Part.from_bytes(data=audio_bytes, mime_type=detected_mime)
    prompt = "Transcreva fielmente este áudio. Retorne só a transcrição."
    # Transcrição também usa GEN_CONFIG para não censurar o que você diz
    resp = client.models.generate_content(model=CHAT_MODEL, contents=[prompt, audio_part], config=GEN_CONFIG)
    return (getattr(resp, "text", "") or "").strip()

# ===== HTML TEMPLATE (Mantenho seu design Gemini White) =====
HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="pt-br">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Aeterna</title>
  <style>
    body { background:#fff; color:#1f1f1f; font-family: sans-serif; margin:0; display:flex; flex-direction:column; height:100vh; overflow:hidden; }
    #chat { flex:1; overflow-y:auto; padding:18px; display:flex; flex-direction:column; gap:14px; padding-bottom:150px; }
    .msg { max-width:86%; padding:12px 16px; border-radius:18px; font-size:1.05rem; line-height:1.35; white-space:pre-wrap; }
    .u { align-self:flex-end; background:#f0f4f9; }
    .a { align-self:flex-start; border:1px solid #eee; background:#fff; }
    .top { padding:14px 18px; border-bottom:1px solid #eee; font-weight:600; display:flex; justify-content:space-between; align-items:center; }
    .small { font-size:.95rem; opacity:.8; cursor:pointer; }
    .area { position:fixed; bottom:18px; left:0; right:0; display:flex; justify-content:center; padding:10px; }
    .cont { width:95%; max-width:900px; background:#f0f4f9; border-radius:28px; display:flex; align-items:center; gap:10px; padding:8px 12px; }
    textarea { flex:1; border:none; background:transparent; outline:none; font-size:1.1rem; padding:10px; resize:none; }
    button { background:none; border:none; font-size:1.3rem; cursor:pointer; color:#333; }
    img.gen { max-width: 340px; border-radius: 14px; border:1px solid #eee; display:block; margin-top:10px; }
  </style>
</head>
<body>
  <div class="top"><div>Aeterna</div><div class="small" onclick="clearMemory()">Limpar memória</div></div>
  <div id="chat"></div>
  <div class="area"><div class="cont">
      <button onclick="document.getElementById('f').click()">➕</button>
      <input type="file" id="f" style="display:none" accept="image/*,audio/*,.pdf,.txt,.docx,.xlsx">
      <textarea id="i" placeholder="Fale comigo..." rows="1"></textarea>
      <button onclick="send()">➤</button>
  </div></div>
  <script>
    const chat = document.getElementById('chat'), input = document.getElementById('i'), fileInput = document.getElementById('f');
    let selectedFile = null;
    const STORAGE_KEY = "aeterna_history_v3";
    function loadHistory() { try { return JSON.parse(localStorage.getItem(STORAGE_KEY) || "[]"); } catch(e){ return []; } }
    function saveHistory(h) { localStorage.setItem(STORAGE_KEY, JSON.stringify(h)); }
    function renderHistory() {
      chat.innerHTML = "";
      loadHistory().forEach(m => {
        if (m.type === "image") chat.innerHTML += `<div class="msg a">${m.text}<br><img class="gen" src="data:${m.mime};base64,${m.b64}" /></div>`;
        else chat.innerHTML += `<div class="msg ${m.role === "user" ? "u" : "a"}">${m.text}</div>`;
      });
      chat.scrollTop = chat.scrollHeight;
    }
    function clearMemory() { localStorage.removeItem(STORAGE_KEY); renderHistory(); }
    renderHistory();
    function talk(t) { const u = new SpeechSynthesisUtterance(t); u.lang="pt-BR"; window.speechSynthesis.speak(u); }
    fileInput.onchange = (e) => { selectedFile = e.target.files?.[0] || null; if (selectedFile) alert("Arquivo: " + selectedFile.name); };
    async function toBase64(file) { return new Promise((resolve, reject) => { const reader = new FileReader(); reader.onload = () => resolve(String(reader.result).split(",")[1]); reader.onerror = reject; reader.readAsDataURL(file); }); }
    async function send() {
      const v = input.value.trim(); if (!v && !selectedFile) return;
      const h = loadHistory(); if (v) h.push({ role:"user", text:v }); else h.push({ role:"user", text:"Enviando arquivo..." });
      saveHistory(h); renderHistory(); input.value = "";
      const payload = { message: v, history: h.slice(-40) };
      if (selectedFile) { payload.file = await toBase64(selectedFile); payload.mime = selectedFile.type; payload.filename = selectedFile.name; }
      try {
        const res = await fetch("/chat", { method:"POST", headers:{ "Content-Type":"application/json" }, body: JSON.stringify(payload) });
        const d = await res.json();
        const nh = loadHistory();
        if (d.image_b64) nh.push({ type:"image", text: d.reply, b64: d.image_b64, mime: d.image_mime });
        else nh.push({ role:"assistant", text: d.reply });
        saveHistory(nh); renderHistory(); if (d.reply) talk(d.reply);
      } catch(e) { alert("Erro de conexão"); } finally { selectedFile = null; fileInput.value = ""; }
    }
  </script>
</body>
</html>
"""

@app.route("/")
def home(): return render_template_string(HTML_TEMPLATE)

@app.route("/chat", methods=["POST"])
def chat():
    try:
        data = request.get_json(force=True) or {}
        msg = (data.get("message") or "").strip()
        history = data.get("history") or []
        file_b64 = data.get("file"); mime = data.get("mime"); filename = data.get("filename")

        lower = msg.lower()
        if lower.startswith("/imgpro "): return jsonify(generate_with_nano_banana(msg[len("/imgpro "):].strip(), True))
        if lower.startswith("/img "): return jsonify(generate_with_nano_banana(msg[len("/img "):].strip(), False))
        if lower.startswith("/imagen "): return jsonify(generate_with_imagen(msg[len("/imagen "):].strip()))

        extra_context = None; image_part = None
        if file_b64:
            raw = base64.b64decode(file_b64)
            if mime and mime.startswith("image/"): image_part = types.Part.from_bytes(data=raw, mime_type=mime)
            elif (mime and mime.startswith("audio/")) or filename.lower().endswith((".ogg", ".opus", ".mp3", ".wav", ".m4a")):
                extra_context = f"[Transcrição de Áudio]: {transcribe_audio(raw, mime, filename)}"
            elif filename.lower().endswith(".pdf"): extra_context = f"[PDF]: {extract_text_from_pdf(raw)}"
            elif filename.lower().endswith(".docx"): extra_context = f"[DOCX]: {extract_text_from_docx(raw)}"
            elif filename.lower().endswith(".xlsx"): extra_context = f"[XLSX]: {extract_text_from_xlsx(raw)}"
            elif filename.lower().endswith(".txt"): extra_context = f"[TXT]: {extract_text_from_txt(raw)}"

        prompt = build_prompt_with_history(msg, history, extra_context)
        contents = [prompt]
        if image_part: contents.append(image_part)
        
        # Chamada de Chat SEM RESTRIÇÃO
        resp = client.models.generate_content(model=CHAT_MODEL, contents=contents, config=GEN_CONFIG)
        
        return jsonify({"reply": getattr(resp, "text", "") or "Estou aqui."})
    except Exception as e:
        return jsonify({"reply": f"Erro técnico: {str(e)}"}), 200

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")))
